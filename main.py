"""
SkinSync v2 — EfficientNet-B3 Skin Analysis + Tier-Based Recommendation Engine
Backend API (FastAPI)

Data source: bahan aktif updated.xlsx (17 sheets)
Model: best_skin_model.pth (EfficientNet-B3, 3-class sigmoid: Acne, Blackheads, Dark Spots)
"""

import os
import io
import cv2
import numpy as np
import torch
import torch.nn as nn
from torchvision import transforms, models
from PIL import Image
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from itertools import combinations
from fastapi.staticfiles import StaticFiles

# ── openpyxl for xlsx parsing ─────────────────────────────────────────────────
import openpyxl

app = FastAPI(title="SkinSync v2 — Tier-Based Skin Analysis API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ══════════════════════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════════════════════
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH  = os.path.join(BASE_DIR, "best_skin_model.pth")
XLSX_PATH   = os.path.join(BASE_DIR, "bahan aktif updated.xlsx")

CLASSES = ["Acne", "Blackheads", "Dark Spots"]

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"[SkinSync v2] Device: {device}")

# ══════════════════════════════════════════════════════════════════════════════
# LOAD MODEL
# ══════════════════════════════════════════════════════════════════════════════
if not os.path.exists(MODEL_PATH):
    raise FileNotFoundError(f"[SkinSync v2] Model file not found at: {MODEL_PATH}")

try:
    model = models.efficientnet_b3(weights=None)
    model.classifier[1] = nn.Linear(model.classifier[1].in_features, len(CLASSES))
    model.load_state_dict(torch.load(MODEL_PATH, map_location=device))
    model.to(device)
    model.eval()
    print("[SkinSync v2] Model loaded successfully.")
except Exception as e:
    raise RuntimeError(f"[SkinSync v2] Failed to load model: {e}")

transform = transforms.Compose([
    transforms.Resize((300, 300)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
])


# ══════════════════════════════════════════════════════════════════════════════
# FACE DETECTOR  —  OpenCV Haar Cascade (frontal + profile face)
# ══════════════════════════════════════════════════════════════════════════════
try:
    _frontal_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    _profile_path = cv2.data.haarcascades + "haarcascade_profileface.xml"
    
    face_cascade = cv2.CascadeClassifier(_frontal_path)
    profile_cascade = cv2.CascadeClassifier(_profile_path)
    
    if face_cascade.empty() or profile_cascade.empty():
        raise RuntimeError("Haar cascade XML tidak ditemukan atau gagal dimuat")
    print("[SkinSync v2] Face detectors (Frontal & Profile) dimuat OK.")
except Exception as _e:
    face_cascade = None
    profile_cascade = None
    print(f"[SkinSync v2] PERINGATAN: face detector tidak tersedia - {_e}")


def merge_rects(rects: list) -> list:
    """
    Mengelompokkan kotak deteksi yang saling tumpang tindih (overlap)
    menggunakan Intersection over Union (IoU) sederhana agar tidak double-count.
    """
    if not rects:
        return []
    # Urutkan berdasarkan luas area terbesar
    rects = sorted(rects, key=lambda r: r[2] * r[3], reverse=True)
    merged = []
    
    for r in rects:
        overlap = False
        rx1, ry1, rx2, ry2 = r[0], r[1], r[0] + r[2], r[1] + r[3]
        for m in merged:
            mx1, my1, mx2, my2 = m[0], m[1], m[0] + m[2], m[1] + m[3]
            # Hitung intersection (irisan)
            ix1 = max(rx1, mx1)
            iy1 = max(ry1, my1)
            ix2 = min(rx2, mx2)
            iy2 = min(ry2, my2)
            
            if ix1 < ix2 and iy1 < iy2:
                inter_area = (ix2 - ix1) * (iy2 - iy1)
                union_area = (r[2] * r[3]) + (m[2] * m[3]) - inter_area
                iou = inter_area / union_area if union_area > 0 else 0
                if iou > 0.3:  # threshold overlap 30%
                    overlap = True
                    break
        if not overlap:
            merged.append(r)
    return merged


def validate_face(image_bytes: bytes, label: str) -> None:
    """
    Validasi bahwa foto mengandung tepat 1 wajah.
    - Tidak ada wajah  → HTTP 422
    - Lebih dari 1 wajah → HTTP 422
    Jika detector tidak tersedia, validasi dilewati (fail-open).
    """
    if face_cascade is None or profile_cascade is None:
        return  # detektor tidak tersedia, lewati

    # Decode gambar
    arr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "invalid_image",
                "photo": label,
                "message": f"File foto '{label}' tidak dapat dibaca. Pastikan file tidak rusak.",
            },
        )

    # Perbesar kontras agar deteksi lebih andal di berbagai kondisi cahaya
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray  = clahe.apply(gray)

    # Deteksi wajah depan
    faces_front = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.05,
        minNeighbors=3,
        minSize=(35, 35),
        flags=cv2.CASCADE_SCALE_IMAGE,
    )
    
    # Deteksi wajah samping (profile)
    faces_profile = profile_cascade.detectMultiScale(
        gray,
        scaleFactor=1.05,
        minNeighbors=3,
        minSize=(35, 35),
        flags=cv2.CASCADE_SCALE_IMAGE,
    )

    # Gabungkan semua hasil deteksi
    all_rects = []
    for (x, y, w, h) in faces_front:
        all_rects.append([int(x), int(y), int(w), int(h)])
    for (x, y, w, h) in faces_profile:
        all_rects.append([int(x), int(y), int(w), int(h)])

    # Gabungkan kotak yang mendeteksi wajah yang sama
    unique_faces = merge_rects(all_rects)
    n = len(unique_faces)

    if n == 0:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "face_not_detected",
                "photo": label,
                "message": (
                    f"Tidak ada wajah yang terdeteksi pada foto {label}. "
                    "Pastikan wajah terlihat jelas, pencahayaan cukup, "
                    "dan kamera menghadap wajah secara langsung."
                ),
            },
        )


def predict_image(image_bytes: bytes) -> list[float]:
    """Return sigmoid confidence [Acne, Blackheads, Dark Spots]."""
    try:
        image  = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        tensor = transform(image).unsqueeze(0).to(device)
        with torch.no_grad():
            logits = model(tensor)
            confs  = torch.sigmoid(logits).cpu().numpy()[0]
        return [float(round(float(c), 4)) for c in confs]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal memproses gambar: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# LOAD XLSX DATA AT STARTUP
# ══════════════════════════════════════════════════════════════════════════════
DB = {}  # global data store


def _cell(v):
    """Normalize cell value: strip strings, keep numbers as-is."""
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v


def _num(v, default=0.0):
    """Convert to float, defaulting to 0.0."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return default


def _bool_ish(v):
    """Convert 1/0/0.5/'yes'/'no'/'caution' → float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    mapping = {"yes": 1.0, "no": 0.0, "caution": 0.5, "safe": 1.0, "unsafe": 0.0,
               "1": 1.0, "0": 0.0, "0.5": 0.5}
    return mapping.get(s, 0.0)


def load_xlsx():
    """Load all relevant sheets from the xlsx into memory."""
    if not os.path.exists(XLSX_PATH):
        print(f"[SkinSync v2] ERROR: XLSX not found at {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # ── 1. master sheet ───────────────────────────────────────────────────
    ws = wb["master"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    master = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        d = dict(zip(header, vals))
        master.append({
            "name":             d.get("Active Ingredient", ""),
            "category":         d.get("Category", ""),
            "mechanism":        d.get("Mechanism / Cara Kerja", ""),
            "Acne":             _bool_ish(d.get("Acne")),
            "Blackheads":       _bool_ish(d.get("Blackheads")),
            "Dark Spots":       _bool_ish(d.get("Dark Spots")),
            "sensitivity":      int(_num(d.get("Skin Sensitivity (1-4)"), 1)),
            "Perempuan":        _bool_ish(d.get("Perempuan")),
            "Laki-laki":        _bool_ish(d.get("Laki-laki")),
            "Hamil":            _bool_ish(d.get("Hamil")),
            "Teen":             _bool_ish(d.get("Teen (<18)")),
            "Young Adult":      _bool_ish(d.get("Young Adult (18-25)")),
            "Adult":            _bool_ish(d.get("Adult (26-40)")),
            "Mature":           _bool_ish(d.get("Mature (41+)")),
            "Kulit Kering":     _bool_ish(d.get("Kulit Kering")),
            "Kulit Kombinasi":  _bool_ish(d.get("Kulit Kombinasi")),
            "Kulit Berminyak":  _bool_ish(d.get("Kulit Berminyak")),
            "Kulit Sensitif":   _bool_ish(d.get("Kulit Sensitif")),
            "catatan":          d.get("Catatan Penting", ""),
        })
    DB["master"] = {m["name"]: m for m in master}
    print(f"[SkinSync v2] master: {len(master)} ingredients loaded")

    # ── 2. recommendation_engine sheet ────────────────────────────────────
    ws = wb["recommendation_engine"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rec_engine = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        d = dict(zip(header, vals))
        rec_engine.append({
            "name":               d.get("Active Ingredient", ""),
            "category":           d.get("Category", ""),
            "primary_concern":    d.get("Primary Concern", ""),
            "secondary_concern":  d.get("Secondary Concern", ""),
            "age_group":          d.get("Age Group", ""),
            "skin_type":          d.get("Skin Type", ""),
            "pregnancy_status":   (d.get("Pregnancy Status") or "safe").lower(),
            "pregnancy_note":     d.get("Pregnancy Note", ""),
            "priority_score":     int(_num(d.get("Priority Score (1-10)"), 5)),
            "min_severity":       (d.get("Minimum Severity") or "none").lower(),
            "contraindicated_if": d.get("Contraindicated If", ""),
            "concentration":      d.get("Recommended Concentration", ""),
            "product_form":       d.get("Product Form", ""),
            "combine_with":       d.get("Combine With", ""),
            "avoid_with":         d.get("Avoid With", ""),
            "frequency":          d.get("Frequency", ""),
            "am_pm":              d.get("AM/PM", ""),
            "mechanism_short":    d.get("Mechanism Short", ""),
            "evidence":           d.get("Source / Evidence Level", ""),
        })
    DB["rec_engine"] = rec_engine
    print(f"[SkinSync v2] recommendation_engine: {len(rec_engine)} entries loaded")

    # ── 3. scoring_logic sheet ────────────────────────────────────────────
    ws = wb["scoring_logic"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    scoring = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        d = dict(zip(header, vals))
        scoring.append({
            "concern":         d.get("Concern", "").upper(),
            "conf_min":        _num(d.get("Confidence Min")),
            "conf_max":        _num(d.get("Confidence Max")),
            "severity_label":  d.get("Severity Label", ""),
            "tier":            d.get("Recommended Tier", ""),
            "top_ingredients": d.get("Top Ingredients (Priority Order)", ""),
            "notes":           d.get("Notes for Engine", ""),
        })
    DB["scoring_logic"] = scoring
    print(f"[SkinSync v2] scoring_logic: {len(scoring)} tiers loaded")

    # ── 4. ingredient_interactions sheet ───────────────────────────────────
    ws = wb["ingredient_interactions"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    interactions = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        d = dict(zip(header, vals))
        interactions.append({
            "a":              d.get("Ingredient A", ""),
            "b":              d.get("Ingredient B", ""),
            "interaction":    d.get("Interaction Type", ""),
            "severity":       (d.get("Severity Level") or "low").lower(),
            "can_combine":    (d.get("Can Combine") or "yes").lower(),
            "recommendation": d.get("Recommendation", ""),
            "notes":          d.get("Notes", ""),
        })
    DB["interactions"] = interactions
    print(f"[SkinSync v2] ingredient_interactions: {len(interactions)} pairs loaded")

    # ── 5. kehamilan sheet ────────────────────────────────────────────────
    ws = wb["kehamilan"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    kehamilan = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        d = dict(zip(header, vals))
        status_text = (d.get("Status / Alasan") or "").lower()
        is_safe = "aman" in status_text
        is_avoid = "hindari" in status_text or "risiko" in status_text or "kontraindikasi" in status_text
        kehamilan.append({
            "kondisi":     d.get("Kondisi", ""),
            "ingredient":  d.get("Bahan Aktif", ""),
            "status_text": d.get("Status / Alasan", ""),
            "is_safe":     is_safe,
            "is_avoid":    is_avoid,
        })
    DB["kehamilan"] = kehamilan
    print(f"[SkinSync v2] kehamilan: {len(kehamilan)} entries loaded")

    # ── 6. age_filter sheet ───────────────────────────────────────────────
    ws = wb["age_filter"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    age_filter = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        age_filter.append(dict(zip(header, vals)))
    DB["age_filter"] = age_filter
    print(f"[SkinSync v2] age_filter: {len(age_filter)} entries loaded")

    # ── 7. skin_type sheet ────────────────────────────────────────────────
    ws = wb["skin_type"]
    header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    skin_type = []
    for row in ws.iter_rows(min_row=2):
        vals = [_cell(c.value) for c in row]
        if not vals[0]:
            continue
        skin_type.append(dict(zip(header, vals)))
    DB["skin_type"] = skin_type
    print(f"[SkinSync v2] skin_type: {len(skin_type)} entries loaded")

    # ── 8. Lifestyle modifier sheets ──────────────────────────────────────
    for sheet_name in ["diet_gula", "diet_tepung", "sunscreen_behavior", "sleep_stress"]:
        ws = wb[sheet_name]
        header = [_cell(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            vals = [_cell(c.value) for c in row]
            if not vals[0]:
                continue
            rows.append(dict(zip(header, vals)))
        DB[sheet_name] = rows
        print(f"[SkinSync v2] {sheet_name}: {len(rows)} entries loaded")

    wb.close()
    print("[SkinSync v2] All data loaded successfully!")


# Load data at startup
load_xlsx()


# ══════════════════════════════════════════════════════════════════════════════
# SCORING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_severity_tier(concern: str, confidence: float) -> dict:
    """Look up scoring_logic to get severity label and tier for a concern."""
    concern_upper = concern.upper()
    # Map names
    concern_map = {
        "ACNE": "ACNE",
        "BLACKHEADS": "BLACKHEADS",
        "DARK SPOTS": "DARK SPOTS / PIGMENTATION",
    }
    lookup = concern_map.get(concern_upper, concern_upper)

    for tier in DB.get("scoring_logic", []):
        if tier["concern"] == lookup and tier["conf_min"] <= confidence <= tier["conf_max"]:
            return {
                "severity": tier["severity_label"],
                "tier": tier["tier"],
                "notes": tier["notes"],
                "top_ingredients": tier["top_ingredients"],
            }
    # Fallback
    return {"severity": "Unknown", "tier": "Unknown", "notes": "", "top_ingredients": ""}


def age_to_group(answer: str) -> str:
    """Map A1 answer to age group key."""
    mapping = {
        "< 18 tahun": "Teen",
        "18–25 tahun": "Young Adult",
        "26–35 tahun": "Adult",
        "36–45 tahun": "Adult",
        "> 45 tahun": "Mature",
    }
    return mapping.get(answer, "Adult")


def skin_answer_to_type(answer: str) -> tuple[str, str]:
    """Map B1 answer to (skin_type_key_for_master, skin_type_label_for_rec_engine)."""
    mapping = {
        "Sangat berminyak & mengkilap di seluruh wajah":                    ("Kulit Berminyak", "Oily"),
        "Berminyak hanya di T-zone (dahi, hidung, dagu), pipi normal":     ("Kulit Kombinasi", "Combination"),
        "Terasa kencang, kering & kadang mengelupas":                       ("Kulit Kering",    "Dry"),
        "Normal, tidak terlalu berminyak/kering":                           ("Kulit Kombinasi", "All"),
        "Tidak menentu, berubah-ubah":                                      ("Kulit Kombinasi", "Combination"),
    }
    return mapping.get(answer, ("Kulit Kombinasi", "All"))


def sensitivity_tolerance(answer: str) -> int:
    """Map B2 answer to max sensitivity level allowed."""
    mapping = {
        "Ya, sangat reaktif — hampir setiap produk baru bisa cocok/tidak": 1,
        "Kadang-kadang — perlu patch test":                                 2,
        "Jarang — kulit saya cukup kuat":                                   3,
        "Tidak pernah mengalami reaksi":                                    4,
    }
    return mapping.get(answer, 4)


def severity_to_rank(s: str) -> int:
    """Convert severity string to numeric rank for comparison."""
    mapping = {
        "none": 0, "no treatment needed": 0, "no treatment": 0, "prevention": 0,
        "mild": 1, "mild (1-5 lesions)": 1, "mild (1-3 spots)": 1,
        "moderate": 2, "moderate (6-20 lesions)": 2, "moderate (several areas)": 2,
        "significant": 3, "severe": 3, "severe (>20 inflamed)": 3,
        "significant (widespread)": 3,
        "very severe": 4, "very severe (cystic)": 4, "severe melasma": 4,
    }
    return mapping.get(s.lower().strip(), 0)


def min_severity_to_rank(s: str) -> int:
    """Convert min_severity from rec_engine to numeric rank."""
    mapping = {"none": 0, "mild": 1, "moderate": 2, "severe": 3}
    return mapping.get(s.lower().strip(), 0)


def concern_matches(primary_concern: str, detected_concerns: list[str]) -> bool:
    """Check if an ingredient's primary concern matches any detected concern."""
    if not primary_concern:
        return False
    pc = primary_concern.lower()
    for concern in detected_concerns:
        cl = concern.lower()
        if cl in pc:
            return True
        # Map special names
        if cl == "dark spots" and ("pih" in pc or "pigment" in pc or "dark" in pc or "bright" in pc or "melasma" in pc):
            return True
        if cl == "acne" and ("acne" in pc or "jerawat" in pc):
            return True
        if cl == "blackheads" and ("blackhead" in pc or "komedo" in pc or "pore" in pc):
            return True
    return False


def age_matches(age_group_str: str, user_age_group: str) -> bool:
    """Check if ingredient's age group matches user's age group."""
    if not age_group_str or age_group_str.lower() == "all" or "all" in age_group_str.lower():
        return True
    ag = age_group_str.lower()
    uag = user_age_group.lower()
    # Handle ranges like "Teen–Adult", "Adult–Mature", "Young Adult–Mature"
    if uag in ag:
        return True
    # Handle "All incl. sensitive", "All incl. teen", "All esp. sensitive"
    if "all" in ag:
        return True
    return False


def skin_type_matches(skin_type_str: str, user_skin_label: str) -> bool:
    """Check if ingredient's skin type matches user's skin type."""
    if not skin_type_str or skin_type_str.lower() == "all":
        return True
    st = skin_type_str.lower()
    ul = user_skin_label.lower()
    if "all" in st:
        return True
    if ul in st:
        return True
    return False


def check_contraindicated(contra_str: str, user_conditions: set) -> bool:
    """Return True if contraindicated (should be EXCLUDED)."""
    if not contra_str:
        return False
    parts = [p.strip().lower() for p in contra_str.split("|")]
    for p in parts:
        if p in user_conditions:
            return True
    return False


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/analyze-skin")
async def analyze_skin(
    # Photos
    file_depan:  UploadFile = File(...),
    file_kiri:   UploadFile = File(...),
    file_kanan:  UploadFile = File(...),
    # Section A
    A1: str = Form("18–25 tahun"),
    A2: str = Form("Laki-laki"),
    # Section B
    B1: str = Form("Normal, tidak terlalu berminyak/kering"),
    B2: str = Form("Jarang — kulit saya cukup kuat"),
    B3: str = Form(""),  # comma-separated multi-select
    # Section C
    C1: str = Form(""),  # comma-separated multi-select
    C2: str = Form("Tidak ada jerawat"),
    C3: str = Form(""),  # comma-separated multi-select
    C5: str = Form("Tidak ada"),
    # Section D
    D1: str = Form("Tidak"),
    D2: str = Form("Tidak"),
    D3: str = Form(""),  # comma-separated
    D4: str = Form(""),  # comma-separated
    D5: str = Form(""),  # free text
    D6: str = Form(""),  # comma-separated
    # Section E
    E1: str = Form("15–60 menit"),
    E2: str = Form("Kadang-kadang"),
    E3: str = Form("7–8 jam per malam (ideal)"),
    E4: str = Form("Sedang — kadang stres tapi terkontrol"),
    E5: str = Form("Seimbang & bervariasi"),
    # Section F
    F1: str = Form(""),  # comma-separated
    F2: str = Form("Sedang (5–10 menit)"),
    F3: str = Form("Rp 100.000 – 300.000"),
    # Section G
    G1: str = Form(""),  # comma-separated
    G2: str = Form(""),  # free text
):
    # ═════════════════════════════════════════════════════════════════════
    # STEP 0: Validasi wajah  —  harus ada tepat 1 wajah di setiap foto
    # ═════════════════════════════════════════════════════════════════════
    b_depan = await file_depan.read()
    b_kiri  = await file_kiri.read()
    b_kanan = await file_kanan.read()

    validate_face(b_depan, "Tampak Depan")
    validate_face(b_kiri,  "Tampak Kiri")
    validate_face(b_kanan, "Tampak Kanan")

    # ═════════════════════════════════════════════════════════════════════
    # STEP 1: Model inference — confidence scores
    # ═════════════════════════════════════════════════════════════════════

    p_depan = predict_image(b_depan)
    p_kiri  = predict_image(b_kiri)
    p_kanan = predict_image(b_kanan)

    # Average across 3 angles
    confidences = {}
    for i, cls in enumerate(CLASSES):
        confidences[cls] = round((p_depan[i] + p_kiri[i] + p_kanan[i]) / 3.0, 4)

    # ═════════════════════════════════════════════════════════════════════
    # STEP 2: scoring_logic — convert confidence to severity + tier
    # ═════════════════════════════════════════════════════════════════════
    face_analysis = {}
    warnings = []
    detected_concerns = []  # concerns with severity > "None"

    for cls in CLASSES:
        conf = confidences[cls]
        if conf >= 0.5:
            st = get_severity_tier(cls, conf)
            face_analysis[cls] = {
                "confidence": conf,
                "severity":   st["severity"],
                "tier":       st["tier"],
            }
            sev_rank = severity_to_rank(st["severity"])
            if sev_rank >= 1:
                detected_concerns.append(cls)
            if sev_rank >= 4 or "referral" in st.get("notes", "").lower() or "Medical" in st.get("tier", ""):
                warnings.append(f"{cls}: Kondisi {st['severity']} terdeteksi.")
        else:
            face_analysis[cls] = {
                "confidence": conf,
                "severity":   "None",
                "tier":       "No treatment needed",
            }

    # If no concerns detected by model, still try to use C1 (user-selected concerns)
    if not detected_concerns and C1:
        c1_parts = [p.strip() for p in C1.split(",") if p.strip()]
        concern_map = {
            "Jerawat aktif (Acne)": "Acne",
            "Komedo (Blackheads/Whiteheads)": "Blackheads",
            "Bekas jerawat / flek hitam (PIH)": "Dark Spots",
            "Flek karena sinar matahari / aging (Hyperpigmentation)": "Dark Spots",
            "Kulit kusam (Dullness)": "Dark Spots",
            "Pori-pori besar (Large Pores)": "Blackheads",
        }
        for c in c1_parts:
            mapped = concern_map.get(c)
            if mapped and mapped not in detected_concerns:
                detected_concerns.append(mapped)

    # ═════════════════════════════════════════════════════════════════════
    # STEP 3: Parse questionnaire answers
    # ═════════════════════════════════════════════════════════════════════
    user_age_group = age_to_group(A1)
    user_gender = A2.strip()
    skin_master_key, skin_rec_label = skin_answer_to_type(B1)
    user_sensitivity_tolerance = sensitivity_tolerance(B2)

    # B3 — diagnosed conditions
    user_diagnoses = set(p.strip().lower() for p in B3.split(",") if p.strip()) if B3 else set()

    # Pregnancy / Breastfeeding
    is_pregnant = D1.strip() != "Tidak"
    is_breastfeeding = D2.strip().lower().startswith("ya")
    is_preg_or_bf = is_pregnant or is_breastfeeding

    # D3 — prescribed meds
    user_meds = set(p.strip().lower() for p in D3.split(",") if p.strip()) if D3 else set()

    # D4 — systemic meds
    user_systemic = set(p.strip().lower() for p in D4.split(",") if p.strip()) if D4 else set()

    # Isotretinoin check
    on_isotretinoin = any("isotretinoin" in m or "accutane" in m or "roaccutane" in m
                         for m in user_systemic)

    # On tretinoin (topical)
    on_tretinoin = any("tretinoin" in m or "adapalene" in m or "retinoid" in m
                       for m in user_meds)

    # D5 — ingredients known to not work
    bad_ingredients = set(p.strip().lower() for p in D5.split(",") if p.strip()) if D5 else set()

    # D6 — allergies
    user_allergies = set(p.strip().lower() for p in D6.split(",") if p.strip()) if D6 else set()

    # Build conditions set for contraindication check
    user_conditions = set()
    if is_pregnant:
        user_conditions.add("pregnant")
    if is_breastfeeding:
        user_conditions.add("breastfeeding")
    if D1.strip() == "Sedang program hamil (TTC)":
        user_conditions.add("ttc")
    if on_tretinoin:
        user_conditions.add("on_tretinoin")
    if on_isotretinoin:
        user_conditions.add("on_isotretinoin")
    if "teen" in user_age_group.lower():
        user_conditions.add("teen")

    # E answers for lifestyle
    user_uv_exposure = E1.strip()
    user_sunscreen = E2.strip()
    user_sleep = E3.strip()
    user_stress = E4.strip()
    user_diet = E5.strip()

    # F1 — existing products
    existing_products = set(p.strip().lower() for p in F1.split(",") if p.strip()) if F1 else set()

    # G1 — primary goals
    user_goals = [p.strip() for p in G1.split(",") if p.strip()] if G1 else []

    # ── Fallback processing if no concerns detected ──
    is_fallback = False
    fallback_ingredients = set()
    if not detected_concerns:
        is_fallback = True
        # 1. Age-based defaults
        age_map = {
            "< 18 tahun": "Teen (<18)",
            "18–25 tahun": "Young Adult (18–25)",
            "26–35 tahun": "Adult (26–40)",
            "36–45 tahun": "Adult (26–40)",
            "> 45 tahun": "Mature (41–55)",
        }
        age_key = age_map.get(A1, "Adult (26–40)")
        age_row = None
        for row in DB.get("age_filter", []):
            if row.get("Kelompok Umur") == age_key:
                age_row = row
                break
        
        # 2. Skin type defaults
        skin_map = {
            "Oily": "Berminyak (Oily)",
            "Dry": "Kering (Dry)",
            "Combination": "Kombinasi (Combination)",
            "All": "Normal",
        }
        skin_key = skin_map.get(skin_rec_label, "Normal")
        skin_row = None
        for row in DB.get("skin_type", []):
            if row.get("Tipe Kulit") == skin_key:
                skin_row = row
                break

        # Parse and combine
        raw_ings = []
        if age_row and age_row.get("Bahan Aktif Prioritas"):
            raw_ings.extend([x.strip().lower() for x in age_row["Bahan Aktif Prioritas"].split(",")])
        if skin_row and skin_row.get("Rekomendasi Bahan Aktif"):
            raw_ings.extend([x.strip().lower() for x in skin_row["Rekomendasi Bahan Aktif"].split(",")])

        # Normalize terms to match database names
        for item in raw_ings:
            import re
            cleaned = re.sub(r'[\d%\-–\.\(\)/]+', '', item).strip()
            if cleaned == "sa" or "salicylic" in cleaned:
                fallback_ingredients.add("salicylic acid")
            elif cleaned == "vitc" or "vit c" in cleaned or "vitamin c" in cleaned:
                fallback_ingredients.add("vitamin c")
                fallback_ingredients.add("ascorbyl")
                fallback_ingredients.add("ascorbic")
            elif "retinol" in cleaned or "tretinoin" in cleaned:
                fallback_ingredients.add("retinol")
                fallback_ingredients.add("tretinoin")
                fallback_ingredients.add("adapalene")
            elif "aha" in cleaned or "lactic" in cleaned or "glycolic" in cleaned or "mandelic" in cleaned:
                fallback_ingredients.add("lactic acid")
                fallback_ingredients.add("glycolic acid")
                fallback_ingredients.add("mandelic acid")
            elif "spf" in cleaned or "sunscreen" in cleaned:
                fallback_ingredients.add("sunscreen")
                fallback_ingredients.add("zinc oxide")
                fallback_ingredients.add("titanium dioxide")
            elif "clay" in cleaned:
                fallback_ingredients.add("kaolin clay")
                fallback_ingredients.add("bentonite clay")
            elif "ha" in cleaned or "hyaluronic" in cleaned:
                fallback_ingredients.add("hyaluronic acid")
            else:
                if cleaned:
                    fallback_ingredients.add(cleaned)

    # ═════════════════════════════════════════════════════════════════════
    # STEP 3b: Kehamilan blacklist
    # ═════════════════════════════════════════════════════════════════════
    pregnancy_blacklist = set()
    pregnancy_warnings = {}
    if is_preg_or_bf:
        for entry in DB.get("kehamilan", []):
            ing_name = (entry.get("ingredient") or "").lower()
            if entry["is_avoid"]:
                pregnancy_blacklist.add(ing_name)
            elif not entry["is_safe"]:
                pregnancy_warnings[ing_name] = entry.get("status_text", "Perlu konsultasi dokter")

    # ═════════════════════════════════════════════════════════════════════
    # STEP 4: Filter recommendation_engine
    # ═════════════════════════════════════════════════════════════════════
    # Get max severity rank per concern for min_severity filtering
    concern_severity_ranks = {}
    for cls in CLASSES:
        sev = face_analysis[cls]["severity"]
        concern_severity_ranks[cls] = severity_to_rank(sev)

    # If isotretinoin, disable all active recommendations
    if on_isotretinoin:
        warnings.append("Anda sedang mengonsumsi isotretinoin — semua bahan aktif kuat dinonaktifkan. Gunakan hanya moisturizer + SPF dasar.")
        return {
            "status": "success",
            "face_analysis": face_analysis,
            "recommendations": [],
            "interactions": [],
            "warnings": warnings,
            "lifestyle_notes": ["Selama isotretinoin: hanya cleanser lembut, moisturizer, dan SPF50+."],
        }

    candidates = []
    for rec in DB.get("rec_engine", []):
        name = rec["name"]
        name_lower = name.lower()

        # Hard filter: pregnancy
        if is_preg_or_bf:
            if rec["pregnancy_status"] == "unsafe":
                continue
            is_blacklisted = False
            for bl_item in pregnancy_blacklist:
                if bl_item in name_lower or name_lower in bl_item:
                    is_blacklisted = True
                    break
            if is_blacklisted:
                continue

        # Hard filter: contraindicated_if
        if check_contraindicated(rec.get("contraindicated_if", ""), user_conditions):
            continue

        # Hard filter: user's known bad ingredients
        if name_lower in bad_ingredients:
            continue

        # Hard filter: existing product overlap check (F1)
        if "retinoid (retinol/tretinoin)" in existing_products or any("retinol" in p or "retinoid" in p for p in existing_products):
            if rec["category"].lower() == "retinoid" or "retinol" in name_lower or "retinoid" in name_lower:
                continue

        # Hard filter: safety allergy check (D6)
        if any("fragrance" in al or "parfum" in al for al in user_allergies):
            notes_text = (rec.get("caution_note", "") + " " + rec.get("pregnancy_note", "") + " " + rec.get("mechanism_short", "")).lower()
            if "fragrance" in notes_text or "parfum" in notes_text or "perfume" in notes_text:
                continue

        # Filter in fallback mode vs normal mode
        if not is_fallback:
            # Filter: concern match
            if not concern_matches(rec["primary_concern"], detected_concerns):
                # Check secondary concern too
                if not concern_matches(rec.get("secondary_concern", ""), detected_concerns):
                    continue

            # Filter: age group
            if not age_matches(rec["age_group"], user_age_group):
                continue

            # Filter: skin type
            if not skin_type_matches(rec["skin_type"], skin_rec_label):
                continue

            # Filter: minimum severity
            rec_min_sev = min_severity_to_rank(rec["min_severity"])
            has_enough_severity = False
            pc = (rec["primary_concern"] or "").lower()
            for cls in detected_concerns:
                if cls.lower() in pc or concern_matches(rec["primary_concern"], [cls]):
                    if concern_severity_ranks.get(cls, 0) >= rec_min_sev:
                        has_enough_severity = True
                        break
            if rec_min_sev > 0 and not has_enough_severity:
                # Also check if secondary concern has enough severity
                sc = (rec.get("secondary_concern") or "").lower()
                for cls in detected_concerns:
                    if cls.lower() in sc or concern_matches(rec.get("secondary_concern", ""), [cls]):
                        if concern_severity_ranks.get(cls, 0) >= rec_min_sev:
                            has_enough_severity = True
                            break
                if not has_enough_severity:
                    continue
        else:
            # Fallback mode matches from default lists
            matched_fallback = False
            for fb_ing in fallback_ingredients:
                if fb_ing in name_lower or name_lower in fb_ing:
                    matched_fallback = True
                    break
            if not matched_fallback:
                continue

            # Filter: age group (secondary validation for fallback)
            if not age_matches(rec["age_group"], user_age_group):
                continue

            # Filter: skin type (secondary validation for fallback)
            if not skin_type_matches(rec["skin_type"], skin_rec_label):
                continue

        candidates.append(rec)

    # ═════════════════════════════════════════════════════════════════════
    # STEP 5: Secondary filter via master sheet
    # ═════════════════════════════════════════════════════════════════════
    filtered = []
    for rec in candidates:
        name = rec["name"]
        master_entry = DB.get("master", {}).get(name)
        if not master_entry:
            filtered.append(rec)
            continue

        # Sensitivity check
        if master_entry["sensitivity"] > user_sensitivity_tolerance:
            continue

        # Gender check
        gender_key = "Perempuan" if user_gender == "Perempuan" else "Laki-laki"
        if master_entry.get(gender_key, 1.0) == 0.0:
            continue

        # Age group check (secondary validation)
        if master_entry.get(user_age_group, 0.5) == 0.0:
            continue

        # Skin type check (secondary validation)
        if master_entry.get(skin_master_key, 0.5) == 0.0:
            continue

        # Check sensitive skin if user is sensitive
        if "sensitif" in skin_master_key.lower() or user_sensitivity_tolerance <= 1:
            if master_entry.get("Kulit Sensitif", 0.5) == 0.0:
                continue

        filtered.append(rec)

    # ═════════════════════════════════════════════════════════════════════
    # STEP 6: Ingredient interactions validation
    # ═════════════════════════════════════════════════════════════════════
    rec_names = [r["name"] for r in filtered]
    interaction_results = []
    to_remove = set()

    for (i, name_a), (j, name_b) in combinations(enumerate(rec_names), 2):
        for ix in DB.get("interactions", []):
            pair_match = (
                (ix["a"].lower() in name_a.lower() or name_a.lower() in ix["a"].lower()) and
                (ix["b"].lower() in name_b.lower() or name_b.lower() in ix["b"].lower())
            ) or (
                (ix["a"].lower() in name_b.lower() or name_b.lower() in ix["a"].lower()) and
                (ix["b"].lower() in name_a.lower() or name_a.lower() in ix["b"].lower())
            )
            if not pair_match:
                continue

            if ix["can_combine"] == "no":
                # Remove the one with lower priority score
                score_a = filtered[i]["priority_score"]
                score_b = filtered[j]["priority_score"]
                remove_idx = j if score_a >= score_b else i
                to_remove.add(remove_idx)
                interaction_results.append({
                    "a": name_a, "b": name_b,
                    "type": "blocked",
                    "severity": ix["severity"],
                    "recommendation": ix["recommendation"],
                    "notes": ix.get("notes", ""),
                })
            elif ix["can_combine"] == "conditional":
                interaction_results.append({
                    "a": name_a, "b": name_b,
                    "type": "conditional",
                    "severity": ix["severity"],
                    "recommendation": ix["recommendation"],
                    "notes": ix.get("notes", ""),
                })
            elif ix["severity"] == "positive":
                interaction_results.append({
                    "a": name_a, "b": name_b,
                    "type": "positive",
                    "severity": "positive",
                    "recommendation": ix["recommendation"],
                    "notes": ix.get("notes", ""),
                })

    # Remove blocked ingredients
    if to_remove:
        filtered = [r for idx, r in enumerate(filtered) if idx not in to_remove]

    # ═════════════════════════════════════════════════════════════════════
    # STEP 7: Lifestyle modifiers — adjust priority score
    # ═════════════════════════════════════════════════════════════════════
    lifestyle_notes = []

    for rec in filtered:
        name_lower = rec["name"].lower()
        master_entry = DB.get("master", {}).get(rec["name"], {})
        adjusted_score = rec["priority_score"]

        # Sleep + Stress modifier
        poor_sleep = "< 5" in user_sleep or "5–6" in user_sleep
        high_stress = "tinggi" in user_stress.lower() or "sangat" in user_stress.lower()
        if poor_sleep or high_stress:
            # Boost barrier repair ingredients
            if any(kw in name_lower for kw in ["ceramide", "centella", "panthenol", "allantoin"]):
                adjusted_score += 1
            # Reduce strong actives
            if any(kw in name_lower for kw in ["retinol", "tretinoin", "glycolic", "aha", "bha"]):
                adjusted_score -= 1

        # Sunscreen behavior
        if user_sunscreen in ["Jarang / tidak pernah", "Kadang-kadang"]:
            # Boost brightening ingredients
            if master_entry.get("Dark Spots", 0) >= 0.5:
                adjusted_score += 0.5

        # E2 - Sunscreen behavior core recommendation (always boost SPF to top if never uses it)
        if user_sunscreen == "Jarang / tidak pernah":
            if rec.get("category", "").lower() == "sunscreen" or any(kw in name_lower for kw in ["zinc oxide", "titanium dioxide", "sunscreen"]):
                adjusted_score += 2.0

        # E1 - UV Exposure sunscreen boost
        if "3–5 jam" in user_uv_exposure or "> 5 jam" in user_uv_exposure:
            if rec.get("category", "").lower() == "sunscreen" or any(kw in name_lower for kw in ["zinc oxide", "titanium dioxide", "sunscreen"]):
                adjusted_score += 1.0

        # Diet modifier
        if "gula" in user_diet.lower() or "manis" in user_diet.lower() or "olahan" in user_diet.lower():
            if master_entry.get("Acne", 0) >= 0.5:
                adjusted_score += 0.5

        if "susu" in user_diet.lower() or "dairy" in user_diet.lower():
            if master_entry.get("Acne", 0) >= 0.5:
                adjusted_score += 0.5

        # E5 - GI diet boost for Niacinamide and antioxidants
        if any(kw in user_diet.lower() for kw in ["gula", "manis", "olahan", "fast food"]):
            if "niacinamide" in name_lower or any(kw in name_lower for kw in ["vitamin c", "vitamin e", "green tea", "resveratrol", "coenzyme q10", "coq10"]):
                adjusted_score += 1.0

        # C3 - Acne area spot treatment vs full-face
        user_acne_areas = set(p.strip().lower() for p in C3.split(",") if p.strip()) if C3 else set()
        has_localized_acne = user_acne_areas and not any(x in user_acne_areas for x in ["menyebar merata", "tidak ada jerawat"])
        if has_localized_acne:
            if "spot" in rec.get("product_form", "").lower() or "gel" in rec.get("product_form", "").lower():
                adjusted_score += 0.5

        rec["_adjusted_score"] = adjusted_score

    # Generate lifestyle notes
    if poor_sleep or high_stress:
        if poor_sleep and high_stress:
            lifestyle_notes.append("Tidur kurang + stres tinggi: skin barrier melemah. Prioritaskan Ceramide, Centella, dan Panthenol.")
        elif poor_sleep:
            lifestyle_notes.append("Kualitas tidur kurang: regenerasi kulit terganggu. Fokus pada barrier repair.")
        else:
            lifestyle_notes.append("Stres tinggi: kortisol meningkat, jerawat & penuaan dipercepat.")

    if user_sunscreen in ["Jarang / tidak pernah"]:
        lifestyle_notes.append("Tidak/jarang pakai sunscreen: risiko hiperpigmentasi & photoaging meningkat. SPF50+ WAJIB!")
    elif user_sunscreen == "Kadang-kadang":
        lifestyle_notes.append("Sunscreen belum rutin: upgrade ke daily SPF 50+ untuk perlindungan optimal.")

    if "gula" in user_diet.lower() or "manis" in user_diet.lower():
        lifestyle_notes.append("Pola makan tinggi gula: glycation merusak kolagen & memperparah jerawat.")
    if "olahan" in user_diet.lower() or "fast food" in user_diet.lower():
        lifestyle_notes.append("Pola makan tinggi makanan olahan: IGF-1 meningkatkan produksi sebum.")

    # Rosacea special handling
    if "rosacea" in user_diagnoses:
        lifestyle_notes.append("Rosacea terdiagnosis: prioritaskan Azelaic Acid + Centella. Hindari eksfoliasi kuat.")
        for rec in filtered:
            if any(kw in rec["name"].lower() for kw in ["azelaic", "centella", "niacinamide"]):
                rec["_adjusted_score"] = rec.get("_adjusted_score", rec["priority_score"]) + 1.5

    # Melasma special handling
    if "melasma" in user_diagnoses:
        lifestyle_notes.append("Melasma terdiagnosis: prioritaskan Tranexamic Acid, Alpha Arbutin, dan VitC.")
        for rec in filtered:
            if any(kw in rec["name"].lower() for kw in ["tranexamic", "arbutin", "vitamin c", "ascorb"]):
                rec["_adjusted_score"] = rec.get("_adjusted_score", rec["priority_score"]) + 1.5

    # G1 tiebreaker boost
    goal_concern_map = {
        "Mengatasi jerawat aktif": "Acne",
        "Memudarkan bekas jerawat & flek": "Dark Spots",
        "Mencerahkan & meratakan warna kulit": "Dark Spots",
        "Menjaga kelembapan kulit": "Hydration",
        "Mengontrol minyak berlebih": "Blackheads",
        "Mengecilkan pori-pori": "Blackheads",
    }
    for goal in user_goals:
        mapped = goal_concern_map.get(goal)
        if mapped:
            for rec in filtered:
                pc = (rec.get("primary_concern") or "").lower()
                if mapped == "Hydration":
                    if "hydration" in pc or "dry" in pc:
                        rec["_adjusted_score"] = rec.get("_adjusted_score", rec["priority_score"]) + 0.5
                else:
                    if mapped.lower() in pc:
                        rec["_adjusted_score"] = rec.get("_adjusted_score", rec["priority_score"]) + 0.5

    # ═════════════════════════════════════════════════════════════════════
    # STEP 8: Sort by adjusted priority score → return top N
    # ═════════════════════════════════════════════════════════════════════
    filtered.sort(key=lambda x: x.get("_adjusted_score", x["priority_score"]), reverse=True)
    max_products = 5
    top_n = filtered[:max_products]

    # Build response
    recommendations = []
    for rec in top_n:
        master_entry = DB.get("master", {}).get(rec["name"], {})
        caution_note = master_entry.get("catatan", "") if master_entry else ""

        preg_note = ""
        if is_preg_or_bf and rec["pregnancy_status"] == "caution":
            preg_note = rec.get("pregnancy_note", "") or "Gunakan dengan hati-hati selama kehamilan."

        recommendations.append({
            "name":              rec["name"],
            "category":          rec["category"],
            "priority_score":    int(max(1, min(10, round(rec.get("_adjusted_score", rec["priority_score"]))))),
            "targets":           [c for c in CLASSES if master_entry.get(c, 0) >= 0.5] if master_entry else [],
            "mechanism":         rec.get("mechanism_short", ""),
            "concentration":     rec.get("concentration", ""),
            "product_form":      rec.get("product_form", ""),
            "frequency":         rec.get("frequency", ""),
            "am_pm":             rec.get("am_pm", ""),
            "combine_with":      rec.get("combine_with", ""),
            "avoid_with":        rec.get("avoid_with", ""),
            "pregnancy_status":  rec["pregnancy_status"],
            "pregnancy_note":    preg_note,
            "caution_note":      caution_note,
            "evidence":          rec.get("evidence", ""),
        })

    return {
        "status":          "success",
        "face_analysis":   face_analysis,
        "recommendations": recommendations,
        "interactions":    interaction_results,
        "warnings":        warnings,
        "lifestyle_notes": lifestyle_notes,
        "is_preg_or_bf":   is_preg_or_bf,
    }


app.mount("/", StaticFiles(directory="skinsync/dist", html=True), name="frontend")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
